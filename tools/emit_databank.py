#!/usr/bin/env python3
"""Regenerate defendable_databank.js from DefendAble_Data_Bank_Week_in_Aviation.xlsx (repo root).
Run after any edit to the workbook: python3 tools/emit_databank.py"""
import openpyxl, json, os
root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
wb = openpyxl.load_workbook(os.path.join(root, 'DefendAble_Data_Bank_Week_in_Aviation.xlsx'))
def rows(ws):
    hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    return [{hdr[c-1]: ws.cell(row=r, column=c).value for c in range(1, len(hdr)+1) if hdr[c-1]}
            for r in range(2, ws.max_row+1)]
fl = wb['FLIGHT LOG']; dl = wb['DISRUPTION LOG']
flights = []
for r in rows(fl):
    flights.append(dict(
        date=r["Date"], dow=r["Day"], fno=r["Flight #"], reg=r["Reg"], actype=r["A/C Type"],
        aoc=r["AOC"], carrier=r["Carrier"], frm=r["From"], to=r["To"], km=r["Dist (km)"],
        band=r["Band (€)"], rot=r["Rotation"], leg=r["Sector"], std=str(r["STD"]).zfill(4),
        atd=str(r["ATD"]).zfill(4) if r["ATD"] not in (None,"—") else "", sta=str(r["STA"]).zfill(4),
        ata=str(r["ATA"]).zfill(4) if r["ATA"] not in (None,"—") else "",
        depDelay=r["Dep delay (min)"] if r["Dep delay (min)"] not in (None,"—") else None,
        arrDelay=r["Arr delay (min)"] if r["Arr delay (min)"] not in (None,"—") else None,
        status=r["Status"], divTo=r["Diverted to"] or "", code=str(r["Delay code (IATA)"] or ""),
        reason=r["Delay reason"] or "", causedBy=r["Caused by"] or "", mass=r["Mass code"] or "",
        pax=r["Pax"], crew=r["Crew"], notes=r["Notes"] or ""))
disruptions = []
for r in rows(dl):
    disruptions.append(dict(
        date=r["Date of flight"], fno=r["Flight #"], frm=r["From"], to=r["To"],
        dtype=r["Disruption Type"], dt=r["Date/Time of disruption"], mass=r["Mass Code"] or "",
        comms=r['"Extra" comms to customers?'], cls=r["Classification"], pax=r["Number of Pax"],
        aoc=r["A/C type (UK, EU or Swiss)"], delayStr=r["Length of delay"], reg=r["Registration"],
        rot=r["Rotation"], arrDelay=r["Arrival delay (min)"] if r["Arrival delay (min)"] not in (None,"—") else None,
        status=r["Status"], divTo=r["Diverted to"] or "", code=str(r["IATA delay code"] or ""),
        causedBy=r["Caused by flight"] or "", band=r["Distance band (€/pax)"],
        exposure=r["Total exposure (€)"], claims=r["Claims received"],
        evidence=r["Evidence available"], care=r["Care provided"], view=r["Preliminary EC261/UK261 view"]))
API = '''
var MASS = { W401: "Storm Ingrid — thunderstorms London/Amsterdam (Wed 15 Jul)",
             S502: "French ATC industrial action (Fri 17 Jul)" };
function flightsFor(reg, date){ return FLIGHTS.filter(function(f){return f.reg===reg && f.date===date;})
  .sort(function(a,b){return a.std<b.std?-1:1;}); }
function findFlight(fno, date){ return FLIGHTS.find(function(f){return f.fno===fno && f.date===date;})||null; }
function rootOf(f){ var cur=f, guard=0;
  while(cur && cur.causedBy && guard++<6){ var nxt=findFlight(cur.causedBy, cur.date); if(!nxt) break; cur=nxt; }
  return cur; }
function buildNarrative(d){
  var f = findFlight(d.fno, d.date); if(!f) return "";
  var lof = flightsFor(f.reg, f.date);
  var L = [];
  L.push(f.fno + " " + f.frm + "-" + f.to + " " + f.date + ", " + f.reg + " (" + f.actype + ", " +
    (f.aoc==="UK"?"UK":f.aoc==="EU"?"EU":"Swiss") + " AOC), rotation " + f.rot +
    ". STD " + f.std + (f.atd?" ATD "+f.atd:"") + ", STA " + f.sta + (f.ata?" ATA "+f.ata:"") +
    (f.status==="CANCELLED" ? " — CANCELLED." :
     f.status==="DIVERTED" ? " — diverted to " + f.divTo + ", arrival delay " + f.arrDelay + " mins." :
     f.arrDelay!=null ? " — arrival delay " + f.arrDelay + " mins." : ".") +
    " Pax " + f.pax + ".");
  L.push("Line of flying " + f.reg + " " + f.date + ": " + lof.map(function(s){
    return s.fno + " " + s.frm + "-" + s.to + " " + s.status +
      (s.arrDelay ? " +" + s.arrDelay + "m" : "") + (s.divTo ? " (div " + s.divTo + ")" : "");
  }).join("; ") + ".");
  var root = rootOf(f);
  if (root && root.fno !== f.fno) {
    L.push("Root cause on " + root.fno + " " + root.frm + "-" + root.to + ": " + root.reason +
      (root.notes ? ". " + root.notes : "") + ". Knock-on to " + f.fno + " via late inbound aircraft (IATA 93).");
  } else {
    L.push(f.reason + (f.notes ? ". " + f.notes : "") + (f.mass && MASS[f.mass] ? ". Mass event " + f.mass + " — " + MASS[f.mass] + "." : "."));
  }
  if (d.evidence) L.push("Evidence held: " + d.evidence + ".");
  if (d.care) L.push("Care provided: " + d.care + ".");
  return L.join(" ");
}
/* Claim-liability extraction — the engine's sort of a jumbled ICC file */
function extractLiable(flightRows){
  var rows = flightRows || FLIGHTS;
  var out = [];
  rows.forEach(function(f){
    var liable = false, why = "";
    if (f.status === "CANCELLED") { liable = true; why = "Cancellation — Art 5 rights engage"; }
    else if (f.status === "DIVERTED" && (f.arrDelay||0) >= 180) { liable = true; why = "Diversion with 3h+ arrival delay"; }
    else if ((f.arrDelay||0) >= 180) { liable = true; why = "Arrival delay ≥ 3h — Sturgeon threshold met"; }
    if (liable) out.push({ flight: f, why: why });
  });
  out.sort(function(a,b){
    var d = a.flight.date.split("/").reverse().join("") .localeCompare(b.flight.date.split("/").reverse().join(""));
    if (d) return d;
    return (b.flight.arrDelay||999) - (a.flight.arrDelay||999);
  });
  return out;
}
function disruptionFor(f){
  return DISRUPTIONS.find(function(d){ return d.fno===f.fno && d.date===f.date; }) ||
    { date:f.date, fno:f.fno, frm:f.frm, to:f.to, dtype:f.reason||f.status, evidence:"", care:"" };
}
return { FLIGHTS: FLIGHTS, DISRUPTIONS: DISRUPTIONS, MASS: MASS,
  flightsFor: flightsFor, findFlight: findFlight, rootOf: rootOf,
  buildNarrative: buildNarrative, extractLiable: extractLiable, disruptionFor: disruptionFor };
'''
js = ("/**\n * DefendAble — Demo Data Bank: 'A Week in Aviation' 13-19 Jul 2026.\n"
      " * Source of truth: DefendAble_Data_Bank_Week_in_Aviation.xlsx (repo root) — rows JUMBLED as ICC files arrive.\n"
      " * Regenerate: python3 tools/emit_databank.py. All data fictional.\n */\n"
      "var DefendAbleDataBank = (function(){\n'use strict';\n"
      "var FLIGHTS = " + json.dumps(flights, ensure_ascii=False) + ";\n"
      "var DISRUPTIONS = " + json.dumps(disruptions, ensure_ascii=False) + ";\n" + API + "})();\n"
      "if (typeof module !== \"undefined\" && module.exports) { module.exports = DefendAbleDataBank; }\n")
open(os.path.join(root, 'defendable_databank.js'), 'w').write(js)
print("emitted", len(js))
